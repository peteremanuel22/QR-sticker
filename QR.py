
# streamlit_app.py
# ----------------
# Cooker Production Code Builder — v10 (Column A→B lookup fix)
#
# Key fix:
# - Database.xlsx is read with openpyxl; we explicitly map Column A (Material) -> Column B (symbol),
#   ignoring headers and odd formatting. This ensures lookups always use column A and return column B.
#
# Features preserved:
# - Path-safe resolution (relative to script folder).
# - Strict material normalization (digits only; handles Arabic numerals; removes float .0).
# - Entry grid via st.data_editor (no paste box).
# - Product types: Gas-cooker (Y=GC), Built-in (Y=BI), CKD (Y=Ck) with separate cumulative serials per (Year, Y).
# - Logic fallback from Logic.xlsx (Capacity, Family, option, Color) to compose x.
# - Barcode lookup from barcode.xlsx (Column A → Column B) shown as "الباركود الدولي".
# - Display hides Month, Year, Y, x, code_length.
# - Tabs per Y (GC, BI, Ck).
# - Clear History and Delete selected rows.
# - Diagnostics sidebar + version-safe rerun helper.

import streamlit as st
import pandas as pd
import datetime
import os
from io import BytesIO
from pathlib import Path
import re
from typing import Optional
from openpyxl import load_workbook

# ---------- Compatibility helper ----------
def _safe_rerun():
    """Rerun Streamlit across versions."""
    if hasattr(st, "rerun"):
        st.rerun()
    elif hasattr(st, "experimental_rerun"):
        st.experimental_rerun()

# ====== Resolve base directory to the folder where this script resides ======
try:
    BASE_DIR = Path(__file__).resolve().parent
except NameError:
    BASE_DIR = Path(os.getcwd()).resolve()

APP_DB_FILE      = BASE_DIR / "Database.xlsx"
APP_LOGIC_FILE   = BASE_DIR / "Logic.xlsx"
APP_HISTORY_FILE = BASE_DIR / "History.xlsx"
APP_BARCODE_FILE = BASE_DIR / "barcode.xlsx"

# ========= Normalization helpers =========

ARABIC_TO_ASCII = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")

def normalize_material(value: str) -> str:
    """
    Normalize a material code to canonical numeric-only ASCII:
    - strip spaces
    - convert Arabic numerals to ASCII
    - remove any non-digit characters
    - '500008870.0' → '500008870'
    """
    if pd.isna(value):
        return ""
    s = str(value).strip().translate(ARABIC_TO_ASCII)
    s = re.sub(r"\s+", "", s)
    digits = re.findall(r"\d", s)
    return "".join(digits)

def normalize_material_series(series: pd.Series) -> pd.Series:
    return series.astype(str).map(normalize_material)

# ========= Generic helpers =========

def save_df_to_excel(df: pd.DataFrame, path: Path, sheet: str = "Sheet1"):
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet)
    Path(path).write_bytes(out.getvalue())

def probe_file_access(path: Path) -> dict:
    """Probe file exists/readable and return diagnostic info."""
    info = {"exists": False, "readable": False, "error": None, "abs_path": str(path.resolve()), "size": 0}
    try:
        if path.exists():
            info["exists"] = True
            info["size"] = path.stat().st_size
            with open(path, "rb") as fh:
                _ = fh.read(1024)
            info["readable"] = True
        else:
            info["exists"] = False
    except Exception as e:
        info["error"] = f"{type(e).__name__}: {e}"
        info["readable"] = False
    return info

# ========= A→B readers (openpyxl) =========

def read_first_sheet_ab(path: Optional[Path] = None, bytes_buf: Optional[bytes] = None) -> pd.DataFrame:
    """
    Read the first sheet and build DataFrame from Column A (material code) and Column B (value/symbol).
    - Skips header rows automatically (non-digit materials become empty after normalization).
    - Last occurrence wins for duplicates.
    """
    if path is None and bytes_buf is None:
        return pd.DataFrame(columns=["material", "symbol"])

    try:
        wb = load_workbook(filename=(BytesIO(bytes_buf) if bytes_buf is not None else str(path)), data_only=True)
    except Exception as e:
        st.error(f"Could not open Excel file: {e}")
        return pd.DataFrame(columns=["material", "symbol"])

    ws = wb.worksheets[0]  # first sheet
    mapping = {}
    for row in ws.iter_rows(values_only=True):
        a_val = row[0] if len(row) > 0 else None  # Column A
        b_val = row[1] if len(row) > 1 else None  # Column B
        a = normalize_material("" if a_val is None else str(a_val))
        b = "" if b_val is None else str(b_val).strip()
        # Skip header/empty A or empty B
        if a == "" or b == "":
            continue
        mapping[a] = b  # last occurrence wins

    if not mapping:
        return pd.DataFrame(columns=["material", "symbol"])

    df = pd.DataFrame({"material": list(mapping.keys()), "symbol": list(mapping.values())})
    # Deduplicate just in case (though last wins already)
    df = df.drop_duplicates(subset=["material"], keep="last")
    return df[["material", "symbol"]]

def read_barcode_ab(path: Optional[Path] = None) -> pd.DataFrame:
    """
    Read barcode.xlsx first sheet Column A (material) → Column B (barcode).
    """
    if path is None or not path.exists():
        return pd.DataFrame(columns=["material", "barcode"])
    try:
        wb = load_workbook(filename=str(path), data_only=True)
    except Exception as e:
        st.error(f"Could not open barcode file: {e}")
        return pd.DataFrame(columns=["material", "barcode"])

    ws = wb.worksheets[0]
    mapping = {}
    for row in ws.iter_rows(values_only=True):
        a_val = row[0] if len(row) > 0 else None
        b_val = row[1] if len(row) > 1 else None
        a = normalize_material("" if a_val is None else str(a_val))
        b = "" if b_val is None else str(b_val).strip()
        if a == "" or b == "":
            continue
        mapping[a] = b

    if not mapping:
        return pd.DataFrame(columns=["material", "barcode"])
    df = pd.DataFrame({"material": list(mapping.keys()), "barcode": list(mapping.values())})
    return df.drop_duplicates(subset=["material"], keep="last")[["material", "barcode"]]

# ========= History file =========

def ensure_history_file() -> pd.DataFrame:
    cols = [
        "CookerName", "Quantity", "MaterialCode", "OrderNumber",
        "Month", "Year", "Y", "x", "symbol", "code34", "code_length",
        "الباركود الدولي", "LastSerial"
    ]
    if not APP_HISTORY_FILE.exists():
        hist = pd.DataFrame(columns=cols)
        save_df_to_excel(hist, APP_HISTORY_FILE, sheet="HISTORY")
        return hist
    try:
        hist = pd.read_excel(APP_HISTORY_FILE, sheet_name="HISTORY", engine="openpyxl")
        for c in cols:
            if c not in hist.columns:
                hist[c] = pd.Series(dtype="object")
        # Normalize stored material codes
        if "MaterialCode" in hist.columns:
            hist["MaterialCode"] = normalize_material_series(hist["MaterialCode"])
        return hist[cols]
    except Exception as e:
        st.warning(f"Couldn't read {APP_HISTORY_FILE.name} ({e}). Creating a new empty history.")
        hist = pd.DataFrame(columns=cols)
        save_df_to_excel(hist, APP_HISTORY_FILE, sheet="HISTORY")
        return hist

# ========= Logic loader (pandas) =========

def load_logic() -> dict:
    if not APP_LOGIC_FILE.exists():
        st.info(f"`{APP_LOGIC_FILE.name}` not found. Please upload it (sheets: Capacity, Family, option, Color).")
        upl = st.file_uploader("Upload Logic.xlsx", type=["xlsx"], key="logic_upl")
        if upl is not None:
            APP_LOGIC_FILE.write_bytes(upl.getbuffer())
            st.success(f"{APP_LOGIC_FILE.name} saved to: {BASE_DIR}")
        else:
            st.stop()

    try:
        xls = pd.ExcelFile(APP_LOGIC_FILE, engine="openpyxl")
    except Exception as e:
        st.error(f"Could not read {APP_LOGIC_FILE.name}: {e}")
        st.stop()

    logic = {}
    for sh in ["Capacity", "Family", "option", "Color"]:
        try:
            df = pd.read_excel(xls, sheet_name=sh, dtype=str)
        except Exception as e:
            st.error(f"Sheet '{sh}' missing/unreadable in {APP_LOGIC_FILE.name}: {e}")
            st.stop()
        df = df.iloc[:, :2].copy()
        df.columns = ["name", "code"]
        df["name"] = df["name"].astype(str).str.strip()
        df["code"] = df["code"].astype(str).str.strip()
        logic[sh.lower()] = df.dropna(how="any")
    return logic

# ========= Database loader (A→B) =========

def load_database_from_disk_or_upload() -> pd.DataFrame:
    """
    Prefer on-disk Database.xlsx (Column A→B). If unreadable, allow upload for session.
    """
    db_info = probe_file_access(APP_DB_FILE)
    fallback_bytes = None

    # Sidebar diagnostics + tools
    with st.sidebar:
        st.subheader("Diagnostics")
        st.caption("Ensure files are in the same folder as this script.")
        st.text(f"Script folder:\n{BASE_DIR}")
        st.write("**Database.xlsx status**")
        st.code(
            f"path: {db_info['abs_path']}\nexists: {db_info['exists']}\nreadable: {db_info['readable']}\nsize: {db_info['size']} bytes\nerror: {db_info['error'] or 'None'}"
        )
        # Folder contents
        try:
            items = "\n".join(sorted(p.name for p in BASE_DIR.iterdir()))
        except Exception as e:
            items = f"(Error listing directory: {e})"
        st.write("**Folder contents**")
        st.code(items or "(empty)")

        col_sa, col_sb = st.columns(2)
        with col_sa:
            if st.button("Rescan"):
                _safe_rerun()
        with col_sb:
            repl = st.file_uploader("Replace Database.xlsx", type=["xlsx"], key="db_replace")
            if repl is not None:
                try:
                    APP_DB_FILE.write_bytes(repl.getbuffer())
                    st.success("Database.xlsx replaced successfully.")
                    _safe_rerun()
                except Exception as e:
                    st.error(f"Replace failed: {e}")

        # Session fallback if unreadable
        if not db_info["readable"]:
            st.warning("On-disk Database.xlsx not readable (locked or permission). Upload a temporary copy for this session:")
            tmp = st.file_uploader("Session Database.xlsx", type=["xlsx"], key="db_session")
            if tmp is not None:
                fallback_bytes = tmp.getvalue()
                st.info("Using uploaded Database.xlsx for this session.")

    # Read via openpyxl A→B
    if fallback_bytes is not None:
        df = read_first_sheet_ab(bytes_buf=fallback_bytes)
    else:
        df = read_first_sheet_ab(path=APP_DB_FILE)

    # Return normalized 2-column df
    if df.empty:
        # Create empty file with 2 columns if missing
        empty = pd.DataFrame(columns=["Material", "symbol"])
        save_df_to_excel(empty, APP_DB_FILE, sheet="DATABASE")
        return pd.DataFrame(columns=["material", "symbol"])
    return df

def save_database(df_material_symbol: pd.DataFrame):
    """
    Save database back to disk with headers 'Material' and 'symbol' (first sheet).
    """
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_material_symbol.rename(columns={"material": "Material", "symbol": "symbol"}).to_excel(
            writer, index=False, sheet_name="DATABASE"
        )
    APP_DB_FILE.write_bytes(out.getvalue())

# ========= Barcode loader (A→B) =========

def load_barcode() -> pd.DataFrame:
    return read_barcode_ab(APP_BARCODE_FILE)

# ========= Logic helpers =========

def build_x_from_logic(sel: dict, maps: dict) -> str:
    """Concatenate codes in order: capacity + family + option + color."""
    c = maps["capacity"].get(sel["capacity"], "")
    f = maps["family"].get(sel["family"], "")
    o = maps["option"].get(sel["option"], "")
    col = maps["color"].get(sel["color"], "")
    return f"{c}{f}{o}{col}"

def recompute_last_serial(df_hist: pd.DataFrame) -> pd.DataFrame:
    """Cumulative LastSerial per (Year, Y). Separate counters for GC, BI, Ck within each year."""
    df_hist = df_hist.copy()
    df_hist["Quantity"] = pd.to_numeric(df_hist["Quantity"], errors="coerce").fillna(0).astype(int)
    df_hist = df_hist.sort_index()
    df_hist["LastSerial"] = df_hist.groupby(["Year", "Y"])["Quantity"].cumsum()
    return df_hist

# ========= App state =========
st.set_page_config(page_title="Cooker Production Code Builder", layout="wide")

# Load files (no caching to avoid stale state while we fix access)
history_df = ensure_history_file()
logic      = load_logic()
barcode_df = load_barcode()
db         = load_database_from_disk_or_upload()  # <-- fixed A→B loader

st.session_state.logic   = logic
st.session_state.barcode = barcode_df
st.session_state.history = history_df
st.session_state.db      = db
if "added_materials" not in st.session_state:
    st.session_state.added_materials = set()

# ========= UI: Header =========
st.title("Cooker Production Code Builder")

left, right = st.columns([2, 1])
with left:
    product_type = st.radio(
        "Product type",
        ["Gas-cooker", "Built-in", "CKD"],
        index=0,
        horizontal=True
    )
    Y = {"Gas-cooker": "GC", "Built-in": "BI", "CKD": "Ck"}[product_type]
with right:
    today = datetime.date.today()
    month = st.selectbox(
        "Production month",
        options=list(range(1, 13)),
        format_func=lambda m: datetime.date(2000, m, 1).strftime("%b"),
        index=today.month - 1
    )
    year = st.number_input("Production year", min_value=2000, max_value=2100, value=today.year, step=1)

st.markdown("---")
st.markdown("**Enter rows in the grid (4 columns)** → `CookerName` | `Quantity` | `SAP Material Code` | `OrderNumber`")

# Entry grid
default_df = pd.DataFrame({
    "CookerName": [""],
    "Quantity": [0],
    "MaterialCode": [""],
    "OrderNumber": [""]
})
df_input = st.data_editor(
    default_df,
    num_rows="dynamic",
    use_container_width=True,
    height=220,
    column_config={
        "CookerName": st.column_config.TextColumn("CookerName"),
        "Quantity": st.column_config.NumberColumn("Quantity", min_value=0, step=1),
        "MaterialCode": st.column_config.TextColumn("SAP Material Code"),
        "OrderNumber": st.column_config.TextColumn("OrderNumber"),
    }
)

# Clean input rows + normalize MaterialCode
df_input = df_input.copy()
df_input["MaterialCode"] = normalize_material_series(df_input["MaterialCode"])
df_input = df_input[df_input["MaterialCode"].str.len() > 0]
df_input["CookerName"] = df_input["CookerName"].astype(str).str.strip()
df_input["OrderNumber"] = df_input["OrderNumber"].astype(str).str.strip()
df_input["Quantity"] = pd.to_numeric(df_input["Quantity"], errors="coerce").fillna(0).astype(int)

if not df_input.empty:
    st.subheader("Input preview")
    st.dataframe(df_input, use_container_width=True, height=200)

    # Logic maps
    logic_maps = {
        cat: dict(zip(logic[cat]["name"], logic[cat]["code"]))
        for cat in ["capacity", "family", "option", "color"]
    }

    # DB mapping (Column A→B)
    db_materials = set(db["material"])
    db_map = dict(zip(db["material"], db["symbol"]))

    # Determine matched/missing codes
    unique_input_materials = df_input["MaterialCode"].unique().tolist()
    missing_codes = [m for m in unique_input_materials if m not in db_materials]
    matched_codes = [m for m in unique_input_materials if m in db_materials]

    with st.expander("Matching diagnostics"):
        st.write(f"Matched materials: {len(matched_codes)}")
        if matched_codes:
            st.code("\n".join(matched_codes[:50]))
        st.write(f"Missing materials: {len(missing_codes)}")
        if missing_codes:
            st.code("\n".join(missing_codes[:50]))
        st.write("Sample of DB materials loaded (A column):")
        st.code("\n".join(list(db_materials)[:50]))

    # Gather selections for missing codes
    selections = {}
    if missing_codes:
        st.info(f"⚠️ {len(missing_codes)} material code(s) are not found in the database. Define logic below (Capacity + Family + option + Color):")
        for mat in missing_codes:
            with st.expander(f"Define logic for MaterialCode: {mat}", expanded=True):
                selections[mat] = {
                    "capacity": st.selectbox("Capacity", logic["capacity"]["name"], key=f"{mat}_cap"),
                    "family":   st.selectbox("Family",   logic["family"]["name"],   key=f"{mat}_fam"),
                    "option":   st.selectbox("Option",   logic["option"]["name"],   key=f"{mat}_opt"),
                    "color":    st.selectbox("Color",    logic["color"]["name"],    key=f"{mat}_col"),
                }

    if st.button("Generate & Save"):
        # 1) Lookup/build x
        df_input["x"] = df_input["MaterialCode"].map(db_map)

        new_pairs = []
        for mat in missing_codes:
            sel = selections.get(mat)
            if not sel:
                st.error(f"Selections missing for material {mat}.")
                st.stop()
            x_new = build_x_from_logic(sel, logic_maps)
            df_input.loc[df_input["MaterialCode"] == mat, "x"] = x_new
            new_pairs.append({"material": mat, "symbol": x_new})
            st.session_state.added_materials.add(mat)

        # 2) Update DB (normalized) & save back (A→B)
        if new_pairs:
            new_df = pd.DataFrame(new_pairs)
            new_df["material"] = normalize_material_series(new_df["material"])
            db_updated = pd.concat([db, new_df], ignore_index=True).drop_duplicates(subset=["material"], keep="last")
            st.session_state.db = db_updated
            save_database(st.session_state.db)  # writes headers 'Material' & 'symbol'
            st.success(f"{APP_DB_FILE.name} updated & saved.")
            # refresh maps
            db_materials = set(st.session_state.db["material"])
            db_map = dict(zip(st.session_state.db["material"], st.session_state.db["symbol"]))

        # 3) symbol = MM + YY + Y
        df_input["Month"] = int(month)
        df_input["Year"] = int(year)
        df_input["Y"] = Y
        df_input["symbol"] = df_input.apply(lambda r: f"{int(r['Month']):02d}{int(r['Year']) % 100:02d}{r['Y']}", axis=1)

        # 4) code34 + length
        df_input["code34"] = df_input.apply(
            lambda r: f"{r['x']}&PO={r['OrderNumber']}&PC={r['MaterialCode']}",
            axis=1
        )
        df_input["code_length"] = df_input["code34"].str.len()

        # 5) Barcode → "الباركود الدولي" (Column A→B)
        bmap = dict(zip(barcode_df["material"], barcode_df["barcode"]))
        df_input["الباركود الدولي"] = df_input["MaterialCode"].map(bmap).fillna("")

        # 6) Update history & save
        hist = pd.concat([st.session_state.history, df_input], ignore_index=True)
        hist = recompute_last_serial(hist)
        st.session_state.history = hist
        save_df_to_excel(st.session_state.history, APP_HISTORY_FILE, sheet="HISTORY")
        st.success(f"{APP_HISTORY_FILE.name} updated & saved at: {BASE_DIR}")

        # 7) Show processed (hide Month, Year, Y, x, code_length)
        st.subheader("Processed rows (this batch)")
        display_cols = ["CookerName", "Quantity", "MaterialCode", "OrderNumber", "symbol", "الباركود الدولي", "code34"]
        st.dataframe(df_input[display_cols], use_container_width=True, height=240)

        # 8) Downloads
        out_db = BytesIO()
        with pd.ExcelWriter(out_db, engine="openpyxl") as writer:
            st.session_state.db.rename(columns={"material":"Material"}).to_excel(writer, index=False, sheet_name="DATABASE")
        out_db.seek(0)
        st.download_button(
            "⬇️ Download updated Database.xlsx",
            out_db,
            file_name="Database_updated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        out_batch = BytesIO()
        with pd.ExcelWriter(out_batch, engine="openpyxl") as writer:
            df_input.to_excel(writer, index=False, sheet_name="BATCH")
        out_batch.seek(0)
        st.download_button(
            "⬇️ Download this batch (Excel)",
            out_batch,
            file_name=f"Batch_{year}_{int(month):02d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

st.markdown("---")
st.header("History (persistent)")

# Tabs per Y
hist = recompute_last_serial(st.session_state.history)
st.session_state.history = hist

tabs = st.tabs(["Gas-cooker (GC)", "Built-in (BI)", "CKD (Ck)"])
labels = [("GC", tabs[0]), ("BI", tabs[1]), ("Ck", tabs[2])]

for y_val, tab in labels:
    with tab:
        subset = hist[hist["Y"] == y_val].copy()
        if subset.empty:
            st.info(f"No entries for {y_val}.")
        else:
            show_cols = ["CookerName", "Quantity", "MaterialCode", "OrderNumber", "symbol", "الباركود الدولي", "code34", "LastSerial"]
            st.dataframe(subset[show_cols], use_container_width=True, height=360)

# ===== History maintenance =====
st.subheader("History maintenance")

# Clear entire history
if st.button("🧹 Clear entire History"):
    empty_hist = pd.DataFrame(columns=[
        "CookerName", "Quantity", "MaterialCode", "OrderNumber",
        "Month", "Year", "Y", "x", "symbol", "code34", "code_length",
        "الباركود الدولي", "LastSerial"
    ])
    st.session_state.history = empty_hist
    save_df_to_excel(st.session_state.history, APP_HISTORY_FILE, sheet="HISTORY")
    st.success("History cleared. History.xlsx reset to empty.")
    _safe_rerun()

# Delete selected rows
hist2 = st.session_state.history
indices = st.multiselect(
    "Select one or more rows to delete (by index below)",
    options=hist2.index.tolist(),
    format_func=lambda i: f"#{i} • Order {hist2.loc[i, 'OrderNumber']} • Mat {hist2.loc[i, 'MaterialCode']} • Qty {hist2.loc[i, 'Quantity']} • Y {hist2.loc[i, 'Y']}"
)
if st.button("Delete selected rows"):
    if indices:
        hist3 = hist2.drop(index=indices).sort_index()
        hist3 = recompute_last_serial(hist3)
        st.session_state.history = hist3
        save_df_to_excel(st.session_state.history, APP_HISTORY_FILE, sheet="HISTORY")
        st.success("Selected entries deleted. History.xlsx updated.")
        _safe_rerun()
    else:
        st.warning("No entries selected to delete.")

# Download History
st.download_button(
    "⬇️ Download History.xlsx",
    data=APP_HISTORY_FILE.read_bytes() if APP_HISTORY_FILE.exists() else b"",
    file_name="History.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)

st.markdown("---")
st.caption(
    "If Database.xlsx is open in Excel, Windows may lock it; close Excel or use the sidebar to upload a session copy. "
    "All paths resolve to the script folder."
)
